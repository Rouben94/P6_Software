import serial
import threading 
import string
from time import sleep
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

serial_port = serial.Serial(port="COM44", baudrate=115200, bytesize=8, timeout=2, stopbits=serial.STOPBITS_ONE)

excel_filename = 'analysis.xlsx'
excel_sheet_title = ("meas " + datetime.today().strftime('%H.%M %d-%m-%Y'))

def bm_start(time):
    serial_command = "benchmark_start {}\n"
    serial_port.write(bytes(serial_command.format(time).encode('Ascii')))

def read():
    iterator = 2
    while(1):
        # Wait until there is data waiting in the serial buffer
        if(serial_port.in_waiting > 0):
            serialString = serial_port.readline().decode('Ascii')
            if "<report>" in serialString:
                bm_list = serialString.replace('\x00', '').split()
                bm_list.pop(0)
                bm_list.insert(0, datetime.now().strftime("%H:%M:%S:%f"))
                excel = load_workbook(filename = excel_filename)
                excel_tab = excel[excel_sheet_title]
                for i, excel_char_elem in enumerate(string.ascii_lowercase[:-16]):
                    excel_tab[excel_char_elem+str(iterator)] = bm_list[i]
                iterator += 1
                excel.save(filename = excel_filename)
                print(bm_list)

def write():
    input("Start with enter: ")
    while(1):
        bm_start(60000)
        sleep(90)

if __name__ == "__main__":
    excel = load_workbook(filename = excel_filename)
    excel_tab = excel.create_sheet(title=excel_sheet_title)
    excel_tab['A1'] = "Measurement Time [H:M:S:US]"
    excel_tab['B1'] = "Message ID [n]"
    excel_tab['C1'] = "Network Time Unacknowledged [us]"
    excel_tab['D1'] = "Network Time Acknowledged [us]"
    excel_tab['E1'] = "Number of Hops [n]"
    excel_tab['F1'] = "RSSI [dB]"
    excel_tab['G1'] = "Source Address [n]"
    excel_tab['H1'] = "Destination Address [n]"
    excel_tab['I1'] = "Group Address [n]"
    excel_tab['J1'] = "Data Size Transmited [Bit]"
    excel.save(filename = excel_filename)

    t_read = threading.Thread(target=read)
    t_write = threading.Thread(target=write)
    
    if serial_port.isOpen(): serial_port.close()
    serial_port.open()

    t_read.start()
    t_write.start()